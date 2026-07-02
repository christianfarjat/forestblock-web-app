#!/usr/bin/env python3
"""
Builder: MJM-FB-TI-FOR-001-V0_AppSheet_Schema_Backend.xlsx

Genera el workbook backend que en Fase 3 se convierte a Google Sheet nativo y
alimenta tanto la app de AppSheet como el Apps Script (appsscript/Dataroom.gs).

Pestañas (los nombres deben coincidir con CONFIG.TABS de Dataroom.gs):
  - Documents   : registro documental (pre-cargado desde config/dataroom.config.json)
  - Roles       : rol -> email -> scope
  - Stages      : catálogo de estados (uno dispara el snapshot)
  - Snapshots   : log de copias inmutables (lo llena el webhook)
  - Audit_Log   : bitácora de eventos (lo llena el Apps Script)

Uso:
    python3 build_appsheet_schema.py
Salida:
    builders/output/MJM-FB-TI-FOR-001-V0_AppSheet_Schema_Backend.xlsx
"""

import json
import os
import re

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
CONFIG_PATH = os.path.join(HERE, "..", "config", "dataroom.config.json")
OUT_DIR = os.path.join(HERE, "output")
OUT_NAME = "MJM-FB-TI-FOR-001-V0_AppSheet_Schema_Backend.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")  # verde bosque
HEADER_FONT = Font(bold=True, color="FFFFFF")

# Catálogo de estados. El que dispara el snapshot está marcado.
STAGES = [
    ("DRAFT", "Draft", 1, "FALSE"),
    ("QAQC", "QA/QC Cross-check", 2, "FALSE"),
    ("APPROVED_FOR_VVB", "Approved for VVB", 3, "TRUE"),
    ("SHARED_WITH_VVB", "Shared with VVB", 4, "FALSE"),
    ("VERIFIED", "Verified", 5, "FALSE"),
]


def load_config():
    with open(CONFIG_PATH, encoding="utf-8") as fh:
        return json.load(fh)


def parse_code(filename):
    """MJM-FB-PR-INF-003-V0_Draft_PDD_ENTE.docx -> code, titulo, tipo, sprint."""
    stem = re.sub(r"\.(docx|xlsx|xlsm|pdf)$", "", filename, flags=re.IGNORECASE)
    code, _, slug = stem.partition("_")
    titulo = slug.replace("_", " ").strip() or stem
    sprint_match = re.search(r"Sprint(\d)", filename)
    if sprint_match:
        sprint = "Sprint " + sprint_match.group(1)
    elif "QAQC" in filename or "QA_QC" in filename:
        sprint = "QA/QC"
    else:
        sprint = ""
    # Tipo a partir del segmento del código (PR-INF, TI-PLA, TI-IT, PR-FOR, ...).
    seg = re.search(r"MJM-FB-([A-Z]{2}-[A-Z]{3})", code)
    tipo_map = {
        "PR-INF": "Informe / PDD",
        "TI-PLA": "Plan técnico",
        "TI-IT": "Instructivo",
        "PR-FOR": "Formulario",
        "TI-FOR": "Formulario (backend)",
    }
    tipo = tipo_map.get(seg.group(1), "Documento") if seg else "Índice"
    return code, titulo, tipo, sprint


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_documents(ws, cfg):
    headers = [
        "doc_id", "codigo", "titulo", "tipo", "sprint", "carpeta",
        "drive_file_id", "version", "estado", "stage", "owner_role", "last_update", "snapshot_url",
    ]
    ws.append(headers)
    doc_n = 0
    for carpeta, files in cfg["fileMapping"].items():
        for filename in files:
            doc_n += 1
            code, titulo, tipo, sprint = parse_code(filename)
            ws.append([
                f"DOC-{doc_n:03d}", code, titulo, tipo, sprint, carpeta,
                "", "V0", "Draft", "DRAFT", "INTERNAL", "", "",
            ])
    style_header(ws, len(headers))
    autosize(ws, [10, 26, 40, 20, 10, 34, 20, 8, 12, 18, 12, 20, 22])

    # Dropdown de stage sobre las claves del catálogo Stages.
    stage_keys = ",".join(s[0] for s in STAGES)
    dv = DataValidation(type="list", formula1=f'"{stage_keys}"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"J2:J{ws.max_row}")


def build_roles(ws, cfg):
    headers = ["role", "email", "scope", "active"]
    ws.append(headers)
    for scope, emails in cfg["roles"].items():
        if scope.startswith("$"):
            continue
        for email in emails:
            ws.append([scope, email, scope, "TRUE"])
    if ws.max_row == 1:  # sin filas de datos
        ws.append(["INTERNAL", "christian.farjat@mjmenergia.com", "INTERNAL", "TRUE"])
    style_header(ws, len(headers))
    autosize(ws, [16, 36, 16, 10])


def build_stages(ws):
    headers = ["stage_key", "label", "order", "triggers_snapshot"]
    ws.append(headers)
    for row in STAGES:
        ws.append(list(row))
    style_header(ws, len(headers))
    autosize(ws, [22, 24, 8, 20])


def build_empty(ws, headers, widths):
    ws.append(headers)
    style_header(ws, len(headers))
    autosize(ws, widths)


def main():
    cfg = load_config()
    wb = Workbook()

    ws_docs = wb.active
    ws_docs.title = "Documents"
    build_documents(ws_docs, cfg)

    build_roles(wb.create_sheet("Roles"), cfg)
    build_stages(wb.create_sheet("Stages"))
    build_empty(
        wb.create_sheet("Snapshots"),
        ["snapshot_id", "source_file_id", "source_name", "snapshot_file_id", "snapshot_url", "created_at", "triggered_by"],
        [24, 20, 40, 20, 24, 22, 16],
    )
    build_empty(
        wb.create_sheet("Audit_Log"),
        ["event_id", "timestamp", "actor", "action", "target", "details"],
        [24, 22, 30, 20, 24, 40],
    )

    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, OUT_NAME)
    wb.save(out_path)
    rows = ws_docs.max_row - 1
    print(f"OK  {out_path}")
    print(f"    Documents pre-cargado con {rows} filas · pestañas: {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
