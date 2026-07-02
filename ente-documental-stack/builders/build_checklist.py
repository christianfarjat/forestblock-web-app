#!/usr/bin/env python3
"""
Builder: MJM-FB-PR-FOR-011-V0_Checklist_Maestro_Documentacion_Verra.xlsx

Checklist maestro de documentación Verra (VCS + CCB) para el PDD de ENTE,
organizado por los 8 sprints del PDD + QA/QC, con la referencia al documento
del stack que cubre cada requisito y la referencia metodológica.

Metodología: parametrizable (`--metodologia`). Default VM0042 (Improved Agricultural
Land Management, cubre manejo de pastoreo + SOC). VM0032 (Sustainable Grasslands) aplica
a pastizales; ojo con estepa NATIVA (ver hoja "Metodologia"). Confirmar con equipo técnico.

Uso:
    python3 build_checklist.py [--metodologia VM0042|VM0032]
Salida:
    builders/output/MJM-FB-PR-FOR-011-V0_Checklist_Maestro_Documentacion_Verra.xlsx
"""

import argparse
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DIR = os.path.join(HERE, "output")
OUT_NAME = "MJM-FB-PR-FOR-011-V0_Checklist_Maestro_Documentacion_Verra.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E3D")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CAT_FILL = PatternFill("solid", fgColor="E7F0EB")
TITLE_FONT = Font(bold=True, size=13, color="1F4E3D")

METODOLOGIAS = {
    "VM0042": "VM0042 — Improved Agricultural Land Management (v2.x)",
    "VM0032": "VM0032 — Methodology for Sustainable Grasslands Management",
}

# (categoria/sprint, documento_ref, [(requisito, referencia_verra, metodologia_ref, obligatorio), ...])
SECTIONS = [
    ("Sprint 1 — Identidad", "MJM-FB-PR-INF-004", [
        ("Título del proyecto, proponente y datos de contacto", "VCS Std §3.1", "—", "Sí"),
        ("Ubicación geográfica y límites (coordenadas / KML / shapefile)", "VCS Std §3.8", "—", "Sí"),
        ("Descripción de la actividad: manejo de pastoreo regenerativo", "VCS Std §3.1", "VM0042 §Applicability", "Sí"),
        ("Fecha de inicio del proyecto y su evidencia", "VCS Std §3.5", "—", "Sí"),
        ("Período de acreditación (crediting period) y tipo", "VCS Std §3.5", "—", "Sí"),
        ("Instrumentos de tenencia / derecho de uso de la tierra", "VCS Std §3.14", "—", "Sí"),
    ]),
    ("Sprint 2 — Elegibilidad", "MJM-FB-PR-INF-005", [
        ("Selección y justificación de la metodología aplicable", "VCS Std §3.2", "VM0042 v2.x", "Sí"),
        ("Cumplimiento de las condiciones de aplicabilidad", "Methodology applicability", "VM0042 §4", "Sí"),
        ("Elegibilidad de tierras e historial de uso (período requerido)", "AFOLU/ALM req.", "VM0042 land eligibility", "Sí"),
        ("Estepa NATIVA: aplicabilidad VM0032 vs. exclusión de nativas en VM0042", "SGM consultation", "VM0032", "Sí"),
        ("Titularidad de los créditos (project ownership)", "VCS Std §3.14", "—", "Sí"),
        ("Ruta de cuantificación: measure-and-model / re-measure / default EF", "Quantification approach", "VM0042 §Quant.", "Sí"),
    ]),
    ("Sprint 3 — Baseline y Adicionalidad", "MJM-FB-PR-INF-006", [
        ("Escenario de línea base identificado y justificado", "VCS Std §3.4", "VM0042 baseline", "Sí"),
        ("Adicionalidad: barreras / prácticas comunes", "VCS Std §3.4", "AFOLU additionality tool", "Sí"),
        ("Regulatory surplus", "VCS Std §3.4", "—", "Sí"),
        ("Análisis de inversión (si aplica)", "Investment analysis", "—", "No"),
        ("Muestreo de SOC de línea base y estratificación", "Sampling design", "VM0042 sampling", "Sí"),
    ]),
    ("Sprint 4 — Cuantificación y Andamiaje", "MJM-FB-PR-INF-007", [
        ("Ecuaciones de reducciones de emisiones y remociones de SOC", "Methodology", "VM0042", "Sí"),
        ("Modelo de proceso calibrado (RothC / CENTURY / DNDC) si ruta model", "Model calibration", "VM0042 model", "Condicional"),
        ("Parámetros fijos ex-ante y fuentes", "Methodology", "VM0042", "Sí"),
        ("Estimación ex-ante de GHG benefits del crediting period", "VCS Std §3.4", "—", "Sí"),
        ("Tratamiento de incertidumbre y deducción", "Uncertainty", "VM0042 uncertainty", "Sí"),
        ("Buffer de no-permanencia (AFOLU Non-Permanence Risk Tool)", "AFOLU NPRT", "—", "Sí"),
        ("Evaluación de fugas (leakage)", "VCS AFOLU Req.", "VM0042 leakage", "Sí"),
    ]),
    ("Sprint 5 — Monitoreo", "MJM-FB-PR-INF-008", [
        ("Plan de monitoreo: parámetros ex-post (SOC, pastoreo, insumos)", "VCS Std §3.6", "VM0042 monitoring", "Sí"),
        ("Diseño de muestreo de suelos ex-post (estratos, densidad, profundidad)", "Sampling design", "VM0042", "Sí"),
        ("Procedimientos QA/QC de medición y laboratorio", "VCS Std §3.6", "—", "Sí"),
        ("Registros de manejo (carga animal, rotación, descanso)", "Activity data", "VM0042 activity data", "Sí"),
        ("Estructura organizativa y roles de monitoreo", "VCS Std §3.6", "—", "Sí"),
        ("Gestión de datos, trazabilidad y control de versiones", "VCS Std §3.6", "—", "Sí"),
    ]),
    ("Sprint 6 — Safeguards y Stakeholders", "MJM-FB-PR-INF-009", [
        ("Consulta a stakeholders y período de comentario público", "VCS Std §3.9", "—", "Sí"),
        ("Mecanismo de quejas (grievance redress)", "VCS Std §3.9", "—", "Sí"),
        ("Evaluación de impactos ambientales y sociales", "VCS Std §3.10", "—", "Sí"),
        ("Salvaguardas (no net harm) y respeto de derechos", "AFOLU / CCB", "—", "Sí"),
        ("Comunidades locales / pueblos indígenas si aplica", "CCB CM", "—", "Condicional"),
    ]),
    ("Sprint 7 — Compliance", "MJM-FB-PR-INF-010", [
        ("Cumplimiento de leyes y regulaciones aplicables", "VCS Std §3.11", "—", "Sí"),
        ("No doble conteo / doble reclamo (double counting)", "VCS Std §3.12", "—", "Sí"),
        ("Interacción con NDC / ajustes correspondientes (si aplica)", "Article 6", "—", "Condicional"),
        ("Permanencia y compromisos a largo plazo (AFOLU)", "AFOLU permanence", "—", "Sí"),
        ("Registro y emisión de VCUs", "VCS Registration", "—", "No"),
    ]),
    ("Sprint 8 — CCB", "MJM-FB-PR-INF-011", [
        ("Beneficios netos de clima (Climate)", "CCB CL", "—", "Sí"),
        ("Beneficios netos para comunidades (Community)", "CCB CM", "—", "Sí"),
        ("Beneficios netos de biodiversidad (Biodiversity)", "CCB B", "—", "Sí"),
        ("Escenarios 'sin proyecto' para Community y Biodiversity", "CCB", "—", "Sí"),
        ("Gold Level (opcional)", "CCB GL", "—", "No"),
    ]),
    ("QA/QC — Cross-check", "MJM-FB-PR-INF-012", [
        ("Consistencia cruzada entre secciones del PDD", "Interno", "—", "Sí"),
        ("Trazabilidad numérica (cuantificación ↔ monitoreo ↔ baseline)", "Interno", "—", "Sí"),
        ("Completitud vs. plantilla VCS PD y checklist de validación VVB", "VVB readiness", "—", "Sí"),
        ("Versionado y control de cambios de todos los entregables", "Interno", "—", "Sí"),
        ("Verificación de referencias metodológicas y de estándar", "Interno", "—", "Sí"),
    ]),
]

HEADERS = [
    "item_id", "categoria", "requisito", "referencia_verra", "metodologia_ref",
    "documento_ref", "obligatorio", "estado", "evidencia", "notas",
]
ESTADOS = ["Pendiente", "En progreso", "Listo", "N/A"]


def style_header(ws, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def build_checklist_sheet(ws, metodologia):
    ws.append(HEADERS)
    n = 0
    for categoria, doc_ref, items in SECTIONS:
        for requisito, ref, met_ref, obligatorio in items:
            n += 1
            ws.append([
                f"CHK-{n:03d}", categoria, requisito, ref, met_ref,
                doc_ref, obligatorio, "Pendiente", "", "",
            ])

    style_header(ws, len(HEADERS))

    # Sombreado por categoría (banda suave cuando cambia el sprint).
    prev, shade = None, False
    for r in range(2, ws.max_row + 1):
        cat = ws.cell(row=r, column=2).value
        if cat != prev:
            shade = not shade
            prev = cat
        if shade:
            for col in range(1, len(HEADERS) + 1):
                ws.cell(row=r, column=col).fill = CAT_FILL

    widths = [10, 28, 50, 24, 26, 20, 12, 14, 28, 28]
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="top")

    dv = DataValidation(type="list", formula1='"' + ",".join(ESTADOS) + '"', allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"H2:H{ws.max_row}")
    return n


def build_metodologia_sheet(ws, metodologia):
    ws["A1"] = "Metodología del proyecto"
    ws["A1"].font = TITLE_FONT
    rows = [
        ("", ""),
        ("Seleccionada", METODOLOGIAS.get(metodologia, metodologia)),
        ("", ""),
        ("VM0042", "Improved Agricultural Land Management. Cubre agricultura regenerativa "
                   "incluyendo manejo de pastoreo y remociones de SOC. Rutas: measure-and-model "
                   "(RothC/CENTURY/DNDC), measure-and-re-measure, y default emission factors."),
        ("VM0032", "Sustainable Grasslands Management. Aplica a pastizales/rangelands. "
                   "En consolidación con VM0042."),
        ("", ""),
        ("⚠ Estepa nativa", "ENTE es estepa patagónica (Río Negro). Verra evalúa que los pastizales "
                            "NATIVOS podrían quedar EXCLUIDOS de VM0042 y cubiertos por VM0032. "
                            "CONFIRMAR la metodología aplicable con el equipo técnico antes de fijar el PDD."),
        ("", ""),
        ("Fuentes", "verra.org (VCS Standard, VM0042, VM0032, consultation SGM)."),
    ]
    r = 2
    for a, b in rows:
        ws.cell(row=r, column=1, value=a).font = Font(bold=True)
        c = ws.cell(row=r, column=2, value=b)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 90


def build_workbook(metodologia="VM0042"):
    """Construye y devuelve (workbook, n_requisitos) sin guardarlo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Checklist"
    n = build_checklist_sheet(ws, metodologia)
    build_metodologia_sheet(wb.create_sheet("Metodologia"), metodologia)
    return wb, n


def main():
    parser = argparse.ArgumentParser(description="Genera el checklist maestro Verra de ENTE.")
    parser.add_argument("--metodologia", default="VM0042", choices=list(METODOLOGIAS.keys()),
                        help="Metodología Verra de referencia (default VM0042).")
    args = parser.parse_args()

    wb, n = build_workbook(args.metodologia)
    os.makedirs(OUT_DIR, exist_ok=True)
    out_path = os.path.join(OUT_DIR, OUT_NAME)
    wb.save(out_path)
    print(f"OK  {out_path}")
    print(f"    {n} requisitos en {len(SECTIONS)} secciones · metodología: {args.metodologia}")


if __name__ == "__main__":
    main()
